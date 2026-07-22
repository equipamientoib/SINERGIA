#!/usr/bin/env python3
"""
build_catalogo.py — Convierte la hoja maestra (Excel) en data/catalogo.json,
que es lo que lee la web. Úsalo cada vez que edites el Excel:

    python3 scripts/build_catalogo.py

Lee  : data/sinergia-hoja-maestra.xlsx   (hojas "Equipos" y "Paquetes")
Genera: data/catalogo.json

Para sincronización EN VIVO (sin volver a generar nada), usa el conector de
Google Sheets en google-apps-script/Codigo.gs (ver docs/GOOGLE-SHEETS.md).
"""
import json, os, sys, hashlib
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "data", "sinergia-hoja-maestra.xlsx")
OUT  = os.path.join(ROOT, "data", "catalogo.json")

GRUPO = {  # nombre visible -> clave interna
    "Analizadores y simuladores": "ansim",
    "Instrumentos de medición": "med",
    "Medidores eléctricos": "elec",
    "Herramientas de apoyo": "apoyo",
}

# Parámetros del modelo de precios (editables aquí o en la hoja "Modelo de precios")
MODELO = {
    "igv": 0.18, "alquiler_pct": 0.30, "horas_efectivas": 7,
    "descuento_hora": 0.10, "descuento_dia": 0.45,
    "instrumentista_dia": 120, "instrumentista_min": 60,
    "mult_semana": 4, "mult_mes": 12, "kit_dia": 40,
    "descuento_combinar": {"2": 0.10, "3": 0.12, "4": 0.15},
}

def rows(ws):
    heads = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in (None, ""):
            continue
        yield dict(zip(heads, row))

def num(v):
    try:
        return round(float(v), 2) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None

def build():
    if not os.path.exists(XLSX):
        sys.exit("No encuentro " + XLSX)
    wb = load_workbook(XLSX, data_only=True)

    equipos = []
    for r in rows(wb["Equipos"]):
        dia = num(r.get("precio_dia"))
        e = {
            "id": r["id"], "cat": r.get("categoria", ""),
            "g": GRUPO.get(r.get("grupo", ""), r.get("grupo", "")),
            "scr": r.get("icono") or "num", "code": r.get("codigo") or "",
            "nom": r.get("nombre", ""), "marca": r.get("marca", ""),
            "tier": r.get("tier", ""), "ficha": r.get("ficha_pdf") or "",
            "desc": r.get("descripcion", ""),
        }
        if r.get("foto"):
            e["photo"] = r["foto"]
        if str(r.get("apoyo", "")).strip().lower() in ("sí", "si", "true", "x", "1"):
            e["apoyo"] = True
        if dia is not None:
            e["dia"] = dia
            e["sem"] = round(dia * MODELO["mult_semana"])
            e["mes"] = round(dia * MODELO["mult_mes"])
        try:
            e["specs"] = json.loads(r.get("specs_json") or "{}")
        except (ValueError, TypeError):
            e["specs"] = {}
        equipos.append(e)

    paquetes = []
    for r in rows(wb["Paquetes"]):
        dia = num(r.get("por_dia")) or 0
        kit = str(r.get("kit_apoyo", "")).strip().lower() in ("sí", "si", "true", "x", "1")
        paquetes.append({
            "id": r["id"], "app": r.get("app", ""), "nivel": r.get("nivel", ""),
            "nom": r.get("nombre", ""),
            "items": [s.strip() for s in str(r.get("incluye_ids", "")).split(",") if s.strip()],
            "kit": ["multimetro", "destornillador-elec", "set-46"] if kit else [],
            "pe": num(r.get("por_equipo")), "ph": num(r.get("por_hora")), "dia": dia,
            "psem": round(dia * MODELO["mult_semana"]), "pmes": round(dia * MODELO["mult_mes"]),
            "eqh": num(r.get("equipos_hora")), "eqd": num(r.get("equipos_dia")),
            "desc": r.get("descripcion", ""),
        })

    proyectos = []
    if "Proyectos" in wb.sheetnames:
        for r in rows(wb["Proyectos"]):
            hitos = []
            for hraw in str(r.get("hitos", "") or "").split("|"):
                t = hraw.strip()
                if not t:
                    continue
                done = t.startswith("✓") or t.lower().startswith("ok ")
                hitos.append({"t": t.lstrip("✓").strip(), "d": done})
            clave = str(r.get("clave", "") or "").strip()
            proyectos.append({
                "id": r["id"], "cliente": r.get("cliente", ""),
                "titulo": r.get("titulo", ""), "servicio": r.get("servicio", ""),
                "fecha": r.get("fecha", ""), "foto": r.get("foto") or "",
                "estado": r.get("estado", "En curso"),
                "avance": int(num(r.get("avance")) or 0),
                "desc": r.get("descripcion", ""), "hitos": hitos,
                # La clave nunca se publica: solo su hash SHA-256
                "clave_hash": hashlib.sha256(clave.encode("utf-8")).hexdigest() if clave else "",
            })

    cat = {"equipos": equipos, "paquetes": paquetes, "proyectos": proyectos, "modelo": MODELO}
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(cat, f, ensure_ascii=False, indent=1)
    print("✓ catalogo.json generado: %d equipos, %d paquetes, %d proyectos" % (len(equipos), len(paquetes), len(proyectos)))

if __name__ == "__main__":
    build()
